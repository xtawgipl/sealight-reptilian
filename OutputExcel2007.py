#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time    : 2018/5/10 22:35
# @Author  : zhangjj

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

wb = load_workbook(filename='haggle.xlsx')#获取一个已经存在的excel文件wb
ws=wb.get_sheet_by_name("Sheet4")#打开该文件wb需要用到的worksheet即ws
wb1=Workbook()#新建立一个工作簿wb1
ewb1=ExcelWriter(workbook=wb1)#新建一个ExcelWriter，用来写wb1
dest_filename=r'result.xlsx'#wb1的名字
ws1=wb1.worksheets[0]#取得wb1的第一个工作表ws1
ws1.title="socialrange"#指定ws1名字为socialrange
#for i in range(1,36):
#ws2=wb.get_sheet_by_name("Sheet5")
for i in range(0,36):#此处遍历列1寻找0-35
    li=[]#给每个数建立一个list表
    for row_num in xrange(0,213824):#表示遍历的行数
        c2=ws.cell(row=row_num,column=2).value#获取列2对应于excel中的C列的数据
        c3=ws.cell(row=row_num,column=3).value#获取列3对应于excel中的D列的数据
        #对于每个i遍历列C,找出C列跟i相同的数据，并将对应的D列的不同数据保存到li中
        if c2==i:
            if c3 in li:
                continue
            else:
                li.append(c3)
        else:
            continue
#print len(li)
    ws1.cell(row=i,column=0).value=i
    ws1.cell(row=i,column=1).value=len(li)
ewb1.save(filename=dest_filename)#保存一定要有，否则不会有结果